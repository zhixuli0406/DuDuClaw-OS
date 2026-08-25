# /// script
# requires-python = ">=3.9"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Extract sheet data from a .xlsx or .csv file.

Usage:
    uv run extract.py <input.xlsx|input.csv> [--format json|md]

JSON: {"sheets": {"Sheet1": [[cell, ...], ...]}}. Cell values keep their
native type (numbers stay numbers). CSV is read as a single "Sheet1".
"""
import argparse
import csv
import json
import os
import sys


def read_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = [list(r) for r in csv.reader(f)]
    return {"Sheet1": rows}


def read_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.stderr.write(
            "openpyxl not installed. Run via `uv run` or `pip install openpyxl`.\n"
        )
        sys.exit(2)
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = {}
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else c for c in row])
        sheets[ws.title] = rows
    return sheets


def md_table(rows):
    if not rows:
        return ""
    header = [str(c) for c in rows[0]]
    out = ["| " + " | ".join(header) + " |",
           "| " + " | ".join("---" for _ in header) + " |"]
    for r in rows[1:]:
        cells = [str(c) for c in r]
        cells = (cells + [""] * len(header))[: len(header)]
        out.append("| " + " | ".join(cells) + " |")
    return "\n".join(out)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("--format", choices=["json", "md"], default="json")
    args = ap.parse_args()

    if not os.path.exists(args.input):
        sys.stderr.write(f"Input not found: {args.input}\n")
        sys.exit(1)

    ext = os.path.splitext(args.input)[1].lower()
    sheets = read_csv(args.input) if ext == ".csv" else read_xlsx(args.input)

    if args.format == "json":
        print(json.dumps({"sheets": sheets}, ensure_ascii=False, indent=2, default=str))
    else:
        parts = []
        for name, rows in sheets.items():
            parts.append(f"## {name}\n\n{md_table(rows)}")
        print("\n\n".join(parts))


if __name__ == "__main__":
    main()
