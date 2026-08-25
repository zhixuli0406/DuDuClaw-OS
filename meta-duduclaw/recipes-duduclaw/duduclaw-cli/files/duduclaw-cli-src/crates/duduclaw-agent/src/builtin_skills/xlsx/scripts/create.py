# /// script
# requires-python = ">=3.9"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Create a .xlsx from a JSON or CSV source file.

Usage:
    uv run create.py data.json --out /abs/out.xlsx   # JSON source
    uv run create.py data.csv  --out /abs/out.xlsx   # CSV source

The source kind is inferred from the input extension (`.csv` → CSV, else JSON),
or forced with `--format json|csv`.

JSON schema: {"sheets": {"SheetName": [[header...], [row...], ...]}}. The first
row of each sheet is bold (treated as the header). CSV becomes a single sheet.
"""
import argparse
import csv
import json
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    sys.stderr.write(
        "openpyxl not installed. Run via `uv run` or `pip install openpyxl`.\n"
    )
    sys.exit(2)


def write_sheet(ws, rows):
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("--out", required=True)
    ap.add_argument("--format", choices=["json", "csv"], default=None,
                    help="source kind; inferred from the input extension if omitted")
    args = ap.parse_args()

    kind = args.format or ("csv" if args.input.lower().endswith(".csv") else "json")
    wb = Workbook()
    try:
        if kind == "csv":
            with open(args.input, newline="", encoding="utf-8-sig") as f:
                rows = [list(r) for r in csv.reader(f)]
            ws = wb.active
            ws.title = "Sheet1"
            write_sheet(ws, rows)
        else:
            with open(args.input, encoding="utf-8") as f:
                spec = json.load(f)
            sheets = spec.get("sheets", {})
            if not sheets:
                sys.stderr.write("JSON has no 'sheets'.\n")
                sys.exit(1)
            first = True
            for name, rows in sheets.items():
                ws = wb.active if first else wb.create_sheet()
                ws.title = str(name)[:31]  # Excel sheet-name limit
                write_sheet(ws, rows)
                first = False
        wb.save(args.out)
    except Exception as e:  # noqa: BLE001
        sys.stderr.write(f"Failed to create xlsx: {e}\n")
        sys.exit(1)

    print(args.out)


if __name__ == "__main__":
    main()
